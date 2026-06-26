#!/usr/bin/env python3
"""Export scoring results to a three-sheet Excel workbook.

Usage:
    python export_excel.py <workspace_dir> [--mapping student-mapping.csv]
                                           [--output grades.xlsx]

Sheets:
    1. Grade Table   — per-student scores, percentile, grade, comment
    2. Statistics    — class-level statistics and distribution
    3. Detail        — full per-dimension scores and reasoning summaries
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Grade mapping
# ---------------------------------------------------------------------------

GRADE_LEVELS = [
    (90, "优", "Excellent"),
    (80, "良", "Good"),
    (70, "中", "Satisfactory"),
    (60, "及格", "Pass"),
    (0, "不及格", "Fail"),
]


def weighted_to_percentile(weighted_total: float) -> int:
    """Map 1-5 weighted total to 40-100 percentile scale."""
    return round((weighted_total - 1) / 4 * 60 + 40)


def percentile_to_grade(percentile: int) -> str:
    """Map percentile score to Chinese grade label."""
    for threshold, label, _ in GRADE_LEVELS:
        if percentile >= threshold:
            return label
    return "不及格"


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class StudentMapping:
    anon_id: str
    student_number: str
    name: str


def load_mapping(path: Path) -> dict[str, StudentMapping]:
    """Load student-mapping.csv → dict keyed by anon_id."""
    mapping: dict[str, StudentMapping] = {}
    if not path.exists():
        logger.warning("No mapping file found at %s — using anonymous IDs", path)
        return mapping
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            anon_id = row.get("anon_id", "").strip()
            if anon_id:
                mapping[anon_id] = StudentMapping(
                    anon_id=anon_id,
                    student_number=row.get("student_number", ""),
                    name=row.get("name", ""),
                )
    logger.info("Loaded %d student mappings", len(mapping))
    return mapping


def load_scores(scores_dir: Path) -> list[dict[str, Any]]:
    """Load all score JSON files from a directory."""
    scores = []
    if not scores_dir.exists():
        logger.error("Scores directory not found: %s", scores_dir)
        return scores
    for path in sorted(scores_dir.glob("*.json")):
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            scores.append(data)
        except (json.JSONDecodeError, OSError) as exc:
            logger.warning("Skipping %s: %s", path.name, exc)
    logger.info("Loaded %d score records", len(scores))
    return scores


# ---------------------------------------------------------------------------
# Sheet 1: Grade Table
# ---------------------------------------------------------------------------

def build_grade_table(
    scores: list[dict[str, Any]],
    mapping: dict[str, StudentMapping],
) -> pd.DataFrame:
    """Build the main grade table DataFrame."""
    rows = []
    for idx, score in enumerate(scores, start=1):
        student_id = score.get("student_id", f"unknown-{idx}")
        student_info = mapping.get(student_id)

        weighted_total = score.get("weighted_total", 0.0)
        percentile = score.get("percentile_score") or weighted_to_percentile(weighted_total)
        grade = percentile_to_grade(percentile)

        # Per-dimension scores
        dim_scores = {}
        for dim in score.get("dimension_scores", []):
            dim_name = dim.get("criterion_name", dim.get("criterion_id", "?"))
            dim_scores[dim_name] = dim.get("score", 0)

        # Comment
        comment = score.get("comment", {})
        if isinstance(comment, dict):
            comment_text = comment.get("full_text", "")
            if not comment_text:
                parts = []
                for section in ("strengths", "weaknesses", "suggestions"):
                    val = comment.get(section, "")
                    if val:
                        parts.append(val)
                comment_text = " ".join(parts)
        else:
            comment_text = str(comment)

        # Truncate comment for summary column
        comment_summary = comment_text[:200] if len(comment_text) > 200 else comment_text

        # Confidence and review
        confidence = score.get("overall_confidence", 0.0)
        review_flag = score.get("review_flag", "")
        if not review_flag and confidence < 0.6:
            review_flag = "待复核"

        # Gate status
        gate_status = score.get("gate_status", {})
        if gate_status.get("all_passed", True):
            gate_label = "通过"
        else:
            failed_gates = [
                g["gate_id"]
                for g in gate_status.get("details", [])
                if not g.get("passed", True)
            ]
            gate_label = f"未通过: {', '.join(failed_gates)}" if failed_gates else "警告"

        row = {
            "序号": idx,
            "学号": student_info.student_number if student_info else student_id,
            "姓名": student_info.name if student_info else "",
            **dim_scores,
            "加权总分": round(weighted_total, 2),
            "百分制": percentile,
            "等级": grade,
            "评语摘要": comment_summary,
            "置信度": round(confidence, 2),
            "复核标记": review_flag,
            "门禁状态": gate_label,
        }
        rows.append(row)

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Sheet 2: Statistics
# ---------------------------------------------------------------------------

def build_statistics(
    df: pd.DataFrame,
    scores: list[dict[str, Any]],
) -> list[tuple[str, Any]]:
    """Compute class-level statistics."""
    stats = []

    stats.append(("总人数", len(df)))
    valid = df[df["门禁状态"] == "通过"]
    stats.append(("有效评分数", len(valid)))

    if not valid.empty:
        stats.append(("百分制均分", round(valid["百分制"].mean(), 1)))
        stats.append(("百分制标准差", round(valid["百分制"].std(), 1)))
        stats.append(("最高分", int(valid["百分制"].max())))
        stats.append(("最低分", int(valid["百分制"].min())))
    else:
        stats.append(("百分制均分", "N/A"))
        stats.append(("百分制标准差", "N/A"))
        stats.append(("最高分", "N/A"))
        stats.append(("最低分", "N/A"))

    # Grade distribution
    stats.append(("", ""))
    stats.append(("等级分布", ""))
    for _, label, _ in GRADE_LEVELS:
        count = len(df[df["等级"] == label])
        stats.append((f"  {label}", count))

    # Per-dimension means
    stats.append(("", ""))
    stats.append(("各维度平均分", ""))
    if scores:
        dim_totals: dict[str, list[float]] = {}
        for score in scores:
            for dim in score.get("dimension_scores", []):
                name = dim.get("criterion_name", dim.get("criterion_id", "?"))
                dim_totals.setdefault(name, []).append(dim.get("score", 0))
        for name, values in dim_totals.items():
            mean_val = sum(values) / len(values) if values else 0
            stats.append((f"  {name}", round(mean_val, 2)))

    # Top deduction reasons
    stats.append(("", ""))
    stats.append(("高频扣分点", ""))
    deduction_counts: dict[str, int] = {}
    for score in scores:
        for dim in score.get("dimension_scores", []):
            if dim.get("score", 5) <= 2:
                reason = dim.get("criterion_name", "unknown")
                deduction_counts[reason] = deduction_counts.get(reason, 0) + 1
    top_deductions = sorted(deduction_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    for reason, count in top_deductions:
        stats.append((f"  {reason}", count))

    # Confidence stats
    stats.append(("", ""))
    stats.append(("置信度统计", ""))
    if "置信度" in df.columns and not df.empty:
        stats.append(("  平均置信度", round(df["置信度"].mean(), 2)))
        low_conf = len(df[df["置信度"] < 0.6])
        stats.append(("  低置信度数量 (<0.6)", low_conf))
        stats.append(("  低置信度比例", f"{low_conf / len(df) * 100:.1f}%"))

    # Gate failure counts
    stats.append(("", ""))
    stats.append(("门禁未通过", ""))
    gate_fail_count = len(df[df["门禁状态"] != "通过"])
    stats.append(("  门禁未通过总数", gate_fail_count))

    return stats


# ---------------------------------------------------------------------------
# Sheet 3: Detail
# ---------------------------------------------------------------------------

def build_detail_table(
    scores: list[dict[str, Any]],
    mapping: dict[str, StudentMapping],
) -> pd.DataFrame:
    """Build detailed per-dimension scoring records."""
    rows = []
    for score in scores:
        student_id = score.get("student_id", "?")
        student_info = mapping.get(student_id)
        display_id = student_info.student_number if student_info else student_id

        for dim in score.get("dimension_scores", []):
            reasoning = dim.get("reasoning", "")
            # Truncate long reasoning for readability
            if len(reasoning) > 500:
                reasoning = reasoning[:497] + "..."

            rows.append({
                "学号": display_id,
                "维度": dim.get("criterion_name", dim.get("criterion_id", "?")),
                "权重": dim.get("weight", 0),
                "分数": dim.get("score", 0),
                "证据": dim.get("evidence", "")[:300],
                "推理摘要": reasoning,
                "改进建议": dim.get("improvement", ""),
                "置信度": dim.get("confidence", 0),
            })

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Excel writing with formatting
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, num_cols: int) -> None:
    """Apply header styling to the first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_column_width(ws, max_width: int = 40) -> None:
    """Set column widths based on content."""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, min(len(val), max_width))
        adjusted = min(max_len + 4, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def write_excel(
    grade_df: pd.DataFrame,
    stats: list[tuple[str, Any]],
    detail_df: pd.DataFrame,
    output_path: Path,
) -> None:
    """Write the three-sheet Excel workbook."""
    wb = Workbook()

    # --- Sheet 1: Grade Table ---
    ws1 = wb.active
    ws1.title = "成绩总表"
    # Headers
    for col_idx, col_name in enumerate(grade_df.columns, start=1):
        ws1.cell(row=1, column=col_idx, value=col_name)
    style_header_row(ws1, len(grade_df.columns))
    # Data
    for row_idx, row in enumerate(grade_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    auto_column_width(ws1)
    ws1.freeze_panes = "A2"

    # --- Sheet 2: Statistics ---
    ws2 = wb.create_sheet("统计摘要")
    ws2.cell(row=1, column=1, value="统计项")
    ws2.cell(row=1, column=2, value="数值")
    style_header_row(ws2, 2)
    for row_idx, (label, value) in enumerate(stats, start=2):
        ws2.cell(row=row_idx, column=1, value=label).border = THIN_BORDER
        ws2.cell(row=row_idx, column=2, value=value).border = THIN_BORDER
    auto_column_width(ws2)

    # --- Sheet 3: Detail ---
    ws3 = wb.create_sheet("详细评分")
    if not detail_df.empty:
        for col_idx, col_name in enumerate(detail_df.columns, start=1):
            ws3.cell(row=1, column=col_idx, value=col_name)
        style_header_row(ws3, len(detail_df.columns))
        for row_idx, row in enumerate(detail_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        auto_column_width(ws3)
        ws3.freeze_panes = "A2"

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Excel saved to %s", output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export homework scoring results to Excel.",
    )
    parser.add_argument(
        "workspace",
        type=Path,
        help="Workspace directory containing scores/ subdirectory",
    )
    parser.add_argument(
        "--mapping",
        type=Path,
        default=None,
        help="Path to student-mapping.csv (default: workspace/student-mapping.csv)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output Excel file path (default: workspace/reports/grades.xlsx)",
    )
    args = parser.parse_args()

    workspace: Path = args.workspace
    scores_dir = workspace / "scores"
    mapping_path = args.mapping or (workspace / "student-mapping.csv")
    output_path = args.output or (workspace / "reports" / "grades.xlsx")

    # Load data
    mapping = load_mapping(mapping_path)
    scores = load_scores(scores_dir)
    if not scores:
        logger.error("No scores found in %s", scores_dir)
        sys.exit(1)

    # Build sheets
    grade_df = build_grade_table(scores, mapping)
    stats = build_statistics(grade_df, scores)
    detail_df = build_detail_table(scores, mapping)

    # Write Excel
    write_excel(grade_df, stats, detail_df, output_path)

    # Summary
    logger.info(
        "Export complete: %d students, output: %s",
        len(grade_df),
        output_path,
    )


if __name__ == "__main__":
    main()
